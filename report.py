"""
Copyright 2022 GlaxoSmithKline Research & Development Limited

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

    https://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
"""

'''
    Functions and classes for generating report and writing to a file.
'''
# XXX: Marker to show Pylint & Pylance are done checking the code.


# Build-in modules.
from collections import Counter
from typing import Any, Dict

# Third party modules.
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, colors

# App modules.
import common_functions as cf


def comp_count_for_rpt(mol_comp_ctr: Counter) -> list:
    '''
        Creates molecule composition count list for output.

        Arguments:
            molecule_ctr {Counter} -- Molecule composition.

        Returns:
            {list} -- Molecule composition count list for output.
    '''
    result_list: list = []
    result_list.append(mol_comp_ctr.get('Carbon', 0))
    result_list.append(mol_comp_ctr.get('Hydrogen', 0))
    result_list.append(mol_comp_ctr.get('Nitrogen', 0))
    result_list.append(mol_comp_ctr.get('Oxygen', 0))
    result_list.append(mol_comp_ctr.get('Sulfur', 0))

    return result_list
    # End of function comp_count_for_rpt()


def to_text(value: Any) -> str:
    '''
        Converts value to string.

        Arguements:
            value {Any} -- Value to convert to string.

        Returns:
            (str) -- Value as string.
    '''
    if value is None:
        return ''
    return str(value)
    # End of function to_text()


def write_to_sheet(workbook: Workbook,
                   sheet_name: str,
                   report_text_list: list) -> None:
    '''
        Writes list to a specified sheet in the workbook.

        Arguments:
            workbook {Workbook} -- Workbook to add sheet to.

            sheet_name {str} -- Name of sheet to create and write to.

            report_text_list {list} -- Header and data to be written to the sheet.

        Returns:
            Nothing
    '''
    sheet = workbook.create_sheet(sheet_name)
    sheet.title = sheet_name

    for row in report_text_list:
        sheet.append(row)

    for column_cells in sheet.columns:       # type: ignore
        length = max(len(to_text(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 5       # type: ignore

    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    center_align = Alignment(horizontal='center')
    white_bold_font = Font(bold=True, color=colors.WHITE)

    num_of_title_cols = len(report_text_list[0])

    for col in range(num_of_title_cols):
        col += 1
        sheet.cell(row=1, column=col).fill = header_fill             # type: ignore
        sheet.cell(row=1, column=col).alignment = center_align       # type: ignore
        sheet.cell(row=1, column=col).font = white_bold_font         # type: ignore
    # End of function write_to_sheet()


def write_results_to_file(params_dict: dict,
                          hc_glycans_comp_dict: Dict[str, Counter],
                          lc_glycans_comp_dict: Dict[str, Counter],
                          is_hetrodimer: bool,
                          chem_mod_dict: dict,
                          intact_chem_mod_dict: dict,
                          halfbody_chem_mod_dict: dict,
                          reduced_comps_dict: dict,
                          intact_comps_dict: dict,
                          halfbody_comps_dict: dict,
                          reduced_glycan_comps_dict: Dict[str, Counter],
                          intact_glycan_comps_dict: Dict[str, Counter],
                          halfbody_glycan_comps_dict: Dict[str, Counter]) -> None:
    '''
        Writes output to an Excel file.

        Arguments:
            params_dict {dict} -- Dictionary of input parameters.

            hc_glycans_comp_dict {Dict[str, Counter]} -- Dictionary of heavy chain glycan compositions.

            lc_glycans_comp_dict {Dict[str, Counter]} -- Dictionary of light chain glycan compositions.

            is_hetrodimer {bool} -- Is a hetrodimer.

            chem_mod_dict {dict} -- Dictionary of chemical modification compositions.

            intact_chem_mod_dict {dict} -- Dictionary of chemical modification compositions for intact.

            halfbody_chem_mod_dict {dict} -- Dictionary of chemical modification compositions for halfbody.

            reduced_comps_dict {dict} -- Dictionary of reduced compositions.

            intact_comps_dict {dict} -- Dictionary of intact compositions.

            halfbody_comps_dict {dict} -- Dictionary of half-body compositions.

            reduced_glycan_comps_dict {Dict[str, Counter]} -- Dictionary of reduced glycan compositions.

            intact_glycans_comps_dict {Dict[str, Counter]} -- Dictionary of intact glycan compositions.

            halfbody_gycan_comps_dict {Dict[str, Counter]} -- Dictionary of half-body glycan compositions.

        Returns:
            Nothing
    '''
    ele_name_mass_dict: Dict[str, Dict[str, float]] = cf.gen_ele_name_mass_dict()

    report_text_list: list = []

    workbook = Workbook()

    # 1. User Input.
    report_text_list = []
    report_text_list.append(['Parameter', 'Value'])
    report_text_list.append(['Heavy Chain 1 Sequence', params_dict["HC-1_Sequence"]])
    report_text_list.append(['Heavy Chain 1 Chemical Mod', params_dict["HC-1_Chem_Mod"]])
    report_text_list.append([''])
    report_text_list.append(['Light Chain 1 Sequence', params_dict["LC-1_Sequence"]])
    report_text_list.append(['Light Chain 1 Chemical Mod', params_dict["LC-1_Chem_Mod"]])
    report_text_list.append([''])
    report_text_list.append(['N-Terminal Cyclization 1', params_dict["Cyclize_1"]])
    report_text_list.append(['C-Terminal Lysine Clipping 1', params_dict["Lys_Clip_1"]])
    report_text_list.append([''])
    report_text_list.append(["----------------------------------------", "-------"])
    report_text_list.append([''])
    report_text_list.append(['Heavy Chain 2 Sequence', params_dict["HC-2_Sequence"]])
    report_text_list.append(['Heavy Chain 2 Chemical Mod', params_dict["HC-2_Chem_Mod"]])
    report_text_list.append([''])
    report_text_list.append(['Light Chain 2 Sequence', params_dict["LC-2_Sequence"]])
    report_text_list.append(['Light Chain 2 Chemical Mod', params_dict["LC-2_Chem_Mod"]])
    report_text_list.append([''])
    report_text_list.append(['N-Terminal Cyclization 2', params_dict["Cyclize_2"]])
    report_text_list.append(['C-Terminal Lysine Clipping 2', params_dict["Lys_Clip_2"]])
    report_text_list.append([''])
    report_text_list.append(["----------------------------------------", "-------"])
    report_text_list.append([''])
    report_text_list.append(['Total Number of Disulfides', int(params_dict["Total_Disulfides"])])
    report_text_list.append([''])
    report_text_list.append(['Unreduced HC Disulfides', int(params_dict["HC_Disulfides"])])
    report_text_list.append(['Unreduced LC Disulfides', int(params_dict["LC_Disulfides"])])
    report_text_list.append([''])
    report_text_list.append(['Light Chain is Glycosylated', params_dict["LC_Glycos"]])
    write_to_sheet(workbook=workbook,
                   sheet_name='User Input',
                   report_text_list=report_text_list)


    # 2. Avgerage Element Masses.
    report_text_list = []
    report_text_list.append(['Element', 'Avg Mass (Da)'])
    report_text_list.append(['Carbon', ele_name_mass_dict["Carbon"]["Mass"]])
    report_text_list.append(['Hydrogen', ele_name_mass_dict["Hydrogen"]["Mass"]])
    report_text_list.append(['Nitrogen', ele_name_mass_dict["Nitrogen"]["Mass"]])
    report_text_list.append(['Oxygen', ele_name_mass_dict["Oxygen"]["Mass"]])
    report_text_list.append(['Sulfur', ele_name_mass_dict["Sulfur"]["Mass"]])
    write_to_sheet(workbook=workbook,
                   sheet_name='Avg Element Mass',
                   report_text_list=report_text_list)


    # 3. Glycan Masses.
    report_text_list = []
    report_text_list.append(['Chain', 'Glycan Name', 'Mass (Da)'])

    for glycan_name, glycan_comp_ctr in hc_glycans_comp_dict.items():
        glycan_mass: float = cf.calc_mass_by_ele_name(mol_comp_ctr=glycan_comp_ctr,
                                                      ele_name_mass_dict=ele_name_mass_dict)
        report_text_list.append(['HC', glycan_name, glycan_mass])

    for glycan_name, glycan_comp_ctr in lc_glycans_comp_dict.items():
        glycan_mass: float = cf.calc_mass_by_ele_name(mol_comp_ctr=glycan_comp_ctr,
                                                      ele_name_mass_dict=ele_name_mass_dict)
        report_text_list.append(['LC', glycan_name, glycan_mass])

    write_to_sheet(workbook=workbook,
                   sheet_name='Glycan Masses',
                   report_text_list=report_text_list)


    # 4. Chemical Modification Masses.
    report_text_list = []
    report_text_list.append(['Chain', 'Formula', 'Mass (Da)'])

    for key, value in chem_mod_dict.items():
        chem_mod: str = str(key)             # Chain
        chem_form: str = value['Formula']    # Formual
        chem_mass: float = value['Mass']     # Mass
        report_text_list.append([chem_mod, chem_form, chem_mass])

    write_to_sheet(workbook=workbook,
                   sheet_name='Chem Mod Masses',
                   report_text_list=report_text_list)


    # 5. Reduced Masses.
    report_text_list = []
    report_text_list.append(['Chain', 'Reduction', 'Chem Mod', 'Carbon', 'Hydrogen', 'Nitrogen', 'Oxygen', 'Sulfur', 'Mass (Da)', 'Note'])

    chain_list: list = ['HC-1', 'HC-2', 'LC-1', 'LC-2']
    reduction_type_list: list = ['part', 'full']

    for chain in chain_list:
        for reduction_type in reduction_type_list:
            output_list: list = []
            output_list.append(chain)                                                                                              # Chain
            output_list.append('Partial' if reduction_type == 'part' else 'Full')                                                  # Reduction
            output_list.append(chem_mod_dict[chain]['Formula'])                                                                    # Chem Mod
            output_list.extend(comp_count_for_rpt(mol_comp_ctr=reduced_comps_dict[chain][reduction_type + '_reduced_comp_ctr']))   # Comp Counts (5)
            mol_mass: float = cf.calc_mass_by_ele_name(mol_comp_ctr=reduced_comps_dict[chain][reduction_type + '_reduced_comp_ctr'],
                                                       ele_name_mass_dict=ele_name_mass_dict)
            chem_mod_mass: float = chem_mod_dict[chain]['Mass']
            total_mass: float = mol_mass + chem_mod_mass
            output_list.append(round(total_mass))                                                                                  # Mass
            output_list.append(reduced_comps_dict[chain]['part_reduced_msg'] if reduction_type == 'part' else '')                  # Note

            report_text_list.append(output_list)

    write_to_sheet(workbook=workbook,
                   sheet_name='Reduced Masses',
                   report_text_list=report_text_list)


    # 6. Intact Masses.
    report_text_list = []
    report_text_list.append(['Chain', 'Chem Mod', 'Carbon', 'Hydrogen', 'Nitrogen', 'Oxygen', 'Sulfur', 'Mass (Da)'])

    for intact_mass_label, intact_mass_comp_ctr in intact_comps_dict.items():
        output_list: list = []
        output_list.append(intact_mass_label)                                               # Chain
        output_list.append(intact_chem_mod_dict[intact_mass_label]['Formulas'])             # Chem Mod
        output_list.extend(comp_count_for_rpt(mol_comp_ctr=intact_mass_comp_ctr))           # Comp Counts (5)
        mol_mass: float = cf.calc_mass_by_ele_name(mol_comp_ctr=intact_mass_comp_ctr,
                                                   ele_name_mass_dict=ele_name_mass_dict)
        chem_mod_mass: float = intact_chem_mod_dict[intact_mass_label]['Mass']
        total_mass: float = mol_mass + chem_mod_mass
        output_list.append(round(total_mass))                                               # Mass
        report_text_list.append(output_list)

    write_to_sheet(workbook=workbook,
                   sheet_name='Intact Masses',
                   report_text_list=report_text_list)


    # 7. Half-Body Masses.
    if is_hetrodimer:
        report_text_list = []
        report_text_list.append(['Chain', 'Chem Mod', 'Carbon', 'Hydrogen', 'Nitrogen', 'Oxygen', 'Sulfur', 'Mass (Da)'])

        for halfbody_mass_label, halfbody_mass_comp_ctr in halfbody_comps_dict.items():
            output_list: list = []
            output_list.append(halfbody_mass_label)                                             # Chain
            output_list.append(halfbody_chem_mod_dict[halfbody_mass_label]['Formulas'])         # Chem Mod
            output_list.extend(comp_count_for_rpt(mol_comp_ctr=halfbody_mass_comp_ctr))         # Comp Counts (5)
            mol_mass: float = cf.calc_mass_by_ele_name(mol_comp_ctr=halfbody_mass_comp_ctr,
                                                       ele_name_mass_dict=ele_name_mass_dict)
            chem_mod_mass: float = halfbody_chem_mod_dict[halfbody_mass_label]['Mass']
            total_mass: float = mol_mass + chem_mod_mass
            output_list.append(round(total_mass))                                               # Mass
            report_text_list.append(output_list)

        write_to_sheet(workbook=workbook,
                    sheet_name='Half-Body Masses',
                    report_text_list=report_text_list)

    # 8. Reduced Glycoforms.
    report_text_list = []
    report_text_list.append(['Chain', 'Glycan', 'Chem Mod', 'Carbon', 'Hydrogen','Nitrogen', 'Oxygen', 'Sulfur', 'Mass (Da)'])

    for key, mol_comp_ctr in reduced_glycan_comps_dict.items():
        output_list: list = []
        parts: list = key.split('_', 1)
        chain: str = parts[0]
        glycan: str = parts[1]
        output_list.append(chain)                                                # Chain
        output_list.append(glycan)                                               # Glycan
        output_list.append(chem_mod_dict[chain]['Formula'])                      # Chem Mod
        output_list.extend(comp_count_for_rpt(mol_comp_ctr=mol_comp_ctr))        # Comp Counts (5)
        mol_mass: float = cf.calc_mass_by_ele_name(mol_comp_ctr=mol_comp_ctr,
                                                   ele_name_mass_dict=ele_name_mass_dict)
        chem_mod_mass: float = chem_mod_dict[chain]['Mass']
        total_mass: float = mol_mass + chem_mod_mass
        output_list.append(round(total_mass))                                    # Mass
        report_text_list.append(output_list)

    write_to_sheet(workbook=workbook,
                sheet_name='Reduced Glyco',
                report_text_list=report_text_list)


    # 9. Intact Glycoforms.
    report_text_list = []
    report_text_list.append(['Chain', 'Glycan', 'Chem Mod', 'Carbon', 'Hydrogen','Nitrogen', 'Oxygen', 'Sulfur', 'Mass (Da)'])

    for key, mol_comp_ctr in intact_glycan_comps_dict.items():
        output_list: list = []
        parts: list = key.split('|', 1)
        chain: str = parts[0]
        glycan: str = parts[1]
        output_list.append(chain)                                                       # Chain
        output_list.append(glycan)                                                      # Glycan
        output_list.append(intact_chem_mod_dict[chain]['Formulas'])                     # Chem Mod
        output_list.extend(comp_count_for_rpt(mol_comp_ctr=mol_comp_ctr))               # Comp Counts (5)
        mol_mass: float = cf.calc_mass_by_ele_name(mol_comp_ctr=mol_comp_ctr,
                                                   ele_name_mass_dict=ele_name_mass_dict)
        chem_mod_mass: float = intact_chem_mod_dict[chain]['Mass']
        total_mass: float = mol_mass + chem_mod_mass
        output_list.append(round(total_mass))                                           # Mass
        report_text_list.append(output_list)

    write_to_sheet(workbook=workbook,
                   sheet_name='Intact Glyco',
                   report_text_list=report_text_list)


    # 10. Half-Body Glycoforms.
    if is_hetrodimer:
        report_text_list = []
        report_text_list.append(['Chain', 'Glycan', 'Chem Mod', 'Carbon', 'Hydrogen','Nitrogen', 'Oxygen', 'Sulfur', 'Mass (Da)'])

        for key, mol_comp_ctr in halfbody_glycan_comps_dict.items():
            output_list: list = []
            parts: list = key.split('|', 1)
            chain: str = parts[0]
            glycan: str = parts[1]
            output_list.append(chain)                                                       # Chain
            output_list.append(glycan)                                                      # Glycan
            output_list.append(halfbody_chem_mod_dict[chain]['Formulas'])                   # Chem Mod
            output_list.extend(comp_count_for_rpt(mol_comp_ctr=mol_comp_ctr))               # Comp Counts (5)
            mol_mass: float = cf.calc_mass_by_ele_name(mol_comp_ctr=mol_comp_ctr,
                                                    ele_name_mass_dict=ele_name_mass_dict)
            chem_mod_mass: float = halfbody_chem_mod_dict[chain]['Mass']
            total_mass: float = mol_mass + chem_mod_mass
            output_list.append(round(total_mass))                                           # Mass
            report_text_list.append(output_list)

        write_to_sheet(workbook=workbook,
                    sheet_name='Half-Body Glyco',
                    report_text_list=report_text_list)


    # 11. References.
    report_text_list = []
    report_text_list.append(['#', 'Reference'])
    report_text_list.append([1, "Berglund, Michael, and Michael E. Wieser. \'Isotopic Compositions of the Elements 2009 (IUPAC Technical Report).\' Pure and Applied Chemistry 83, no. 2 (2011): 397-410. https://doi.org/10.1351/PAC-REP-10-06-02."])
    report_text_list.append([2, "Wang, M., G. Audi, A. H. Wapstra, F. G. Kondev, M. MacCormick, X. Xu, and B. Pfeiffer. \'The Ame2012 Atomic Mass Evaluation.\' Chinese Physics C 36, no. 12 (December 2012): 1603-2014. https://doi.org/10.1088/1674-1137/36/12/003."])
    write_to_sheet(workbook=workbook,
                   sheet_name='References',
                   report_text_list=report_text_list)


    workbook.remove(worksheet=workbook["Sheet"])      # type: ignore
    workbook.save(params_dict['File_Path'])
    workbook.close()
    # End of function write_results_to_file()
